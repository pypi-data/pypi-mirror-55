from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from exex import util


class Extractor:
    __book = None

    def __init__(self, excel_file_path):
        self.__book = load_workbook(filename=excel_file_path)

    @property
    def book(self):
        return self.__book

    def __sheet(self):
        first_sheet_name = self.__book.sheetnames[0]
        sheet: Worksheet = self.__book[first_sheet_name]
        return sheet

    def all(self):
        sheet = self.__sheet()
        rows = sheet.values
        result = [list([cell for cell in row]) for row in rows]
        return result

    def range(self, cell_range: any = None):
        sheet = self.__sheet()
        rows = sheet[cell_range]

        if util.is_iterable(rows):
            result = [[cell.value for cell in row] for row in rows]
        else:
            result = [[rows.value]]

        return result

    # TODO
    def cell(self, cell_coord: str):
        pass

    def cells(self, *args):
        return [self.cell(cell_coord) for cell_coord in args]

    # TODO
    def column(self, column_name: str):
        pass

    # TODO
    def columns(self, *args):
        pass

    # TODO
    def row(self, row_number):
        pass

    # TODO
    def rows(self, *args):
        pass
