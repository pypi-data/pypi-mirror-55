from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os,time,datetime,random
import openpyxl,tempfile
from PythonZyjTools.Office.execl_read import xlsx_get_json
from PythonZyjTools import Office
class execl_create:
    # 传入的数据
    data=[]
    def __init__(self):
        self.wb = Workbook()
    # 创建保存文件的路径名
    def create_file_name(self,file_name='',dirs=''):
        """
        创建文件名 创建文件夹
        :param file_name: 文件名
        :param dirs: 目录
        :return: 对象
        """
        dest_filename = ''
        if dirs == '' or type(dirs) != type(''):
            dirs = os.path.dirname(Office.__file__) + '/cache/'
        if file_name == '' or type(file_name) != type(''):
            file_name = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S-") + str(random.randint(1, 300)) + '.xlsx'
            dest_filename = dirs+file_name
            while os.path.exists(dest_filename):
                dest_filename = dirs+file_name
        if os.path.exists(os.path.dirname(dest_filename)) !=True:
            os.makedirs(os.path.dirname(dest_filename))
        return dest_filename
    def jiexi_data(self,data,datatype):
        """
        数据解析器
        :param data: 数据
        :param datatype: 数据类型将以什么格式进行解析
        :return: 返回解析的数据
        """
        if datatype == 'jiegou':
            contecxt = []
            for i in range(len(data)):
                if i == 0:
                    data_hang = []
                    data_hang2 = []
                    for j in data[i] :
                        data_hang.append(list(dict(j).keys())[0])
                        data_hang2.append(list(dict(j).values())[0])
                    contecxt.append(data_hang)
                    contecxt.append(data_hang2)
                else:
                    data_hang = []
                    for j in data[i] :
                        data_hang.append(list(dict(j).values())[0])
                    contecxt.append(data_hang)
            return contecxt
        elif datatype == 'yuanshi':
            return data
        else:
            raise RuntimeError("不能解析传入execl里的数据结构")
    # 加载数据
    def load_data(self,data=[],sheet_title="",is_index=False,datatype="jiegou"):
        """
        加载数据创建文件
        :param data: 载入的数据[[{'a':'1','b':'2'}],[{'a':'1','b':'2'}]]示例两行数据
        :param sheet_title: execl sheet 页面名
        :param is_index: 是否是默认页面
        :param datatype: 数据类型 ，好几种格式交给self.jiexi_data去处理
        :return: 返回self
        """
        if sheet_title != "" and is_index == False:
            self.create_sheet(resource=self.wb,sheet_title=sheet_title,data=self.jiexi_data(data,datatype))
        else:
            self.Index_Page_active(resource=self.wb,data=self.jiexi_data(data,datatype),sheet_title=sheet_title)
        return self
    # 创建sheet页
    def create_sheet(self,resource=None,sheet_title=None,data=None):
        """
        创建sheet页的资源
        :param resource:创建execl的资源
        :param sheet_title:
        :param data:数据 格式为[[一行],[一行]]
        :return:
        """
        if sheet_title != None:
            biaodan = resource.create_sheet(title=sheet_title)
        else:
            biaodan = resource.create_sheet()
        for row in data:
            biaodan.append(row)
        return biaodan
    def Index_Page_active(self,resource=None,data=None,sheet_title=""):
        """
        默认页面的操作
        :param resource: execl资源头
        :param data: 数据
        :return: 资源头
        """
        ws1 = resource.active
        if sheet_title != "" and type(sheet_title) == type(""):
            ws1.title =sheet_title
        for row in data:
            ws1.append(row)
        return ws1
    def save_path(self,path=""):
        """
        创建execl文件
        :param path: 设置保存路径
        :return: 返回文件地址或者是None None表示没有生成成功
        """
        try:
            if path == "":
                path = self.create_file_name()
                self.wb.save(path)
            else:
                self.wb.save(path)
        except:
            return None
        finally:
            self.close()
        return path
    def close(self):
        """
        销毁所有资源
        :return:
        """
        if self.wb != None:
            self.wb.close()
            self.wb = None
        return self
    def zhuijia(self,file_path="",data=[],sheet_name="",title=0,dataType="jiegou"):
        """
        向文件中追加文件信息
        :param file_path: 文件的路径
        :param data: 数据文件
        :param sheet_name: 哪一个sheet页为None时为主页
        :param title:哪一行为标题行
        :return: 返回追加状态
        :param dataType: 数据类型

        """
        try:
            data = self.paixu(xlsx_get_json(file_path).get_data().data,data)

            data = self.jiexi_data(data, dataType)
            xlsx_data = xlsx_get_json(file_path).get_data().data
            xlsx_data = self.jiexi_data(xlsx_data,"jiegou")
            xlsx_data.extend(data)
            self.wb.close()
            wb = openpyxl.load_workbook(file_path)
            self.wb = wb
            self.load_data(data=xlsx_data,is_index=True,datatype='yuanshi',sheet_title=sheet_name)
            self.save_path(file_path)
            return True
        except Exception as question:
            print(question)
            return False
        finally:
            self.close()
    def paixu(self,data,adddata):
        for i in range(data):
            for k in range(adddata):
                pass

