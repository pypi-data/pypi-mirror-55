from .utils.tea_utils import *
from .utils.tea_filter import *
from .utils.auto_bin_woe import AutoBinWOE
from .utils.tea_encoder import *
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold
from openpyxl.chart import BarChart, Reference, LineChart
import warnings
import configparser
import os
current_file_dir = os.path.dirname(__file__)
conf = configparser.ConfigParser()

conf.read(current_file_dir + '/conf.ini', encoding='utf-8')
warnings.filterwarnings('ignore')


class WOE:
    def __init__(self, bins=10, psi_threshold=None, monotony_merge=True, bad_rate_merge=False,
                 bad_rate_sim_threshold=0.05, chi2_merge=False, chi2_threshold=3.841, iv_threshold=None,
                 prune=False, prune_threshold=0.05):
        self.bins = bins
        self.psi_threshold = psi_threshold
        self.bad_rate_merge = bad_rate_merge
        self.bad_rate_sim_threshold = bad_rate_sim_threshold
        self.chi2_merge = chi2_merge
        self.chi2_threshold = chi2_threshold
        self.monotony_merge = monotony_merge
        self.iv_threshold = iv_threshold
        self.prune = prune
        self.prune_threshold = prune_threshold

    def woe_processing(self, x_train, y_train, x_oot, y_oot, gating=True):
        # WOE编码
        woe = AutoBinWOE(bins=self.bins, monotony_merge=self.monotony_merge, bad_rate_merge=self.bad_rate_merge,
                         bad_rate_sim_threshold=self.bad_rate_sim_threshold,
                         chi2_merge=self.chi2_merge, chi2_threshold=self.chi2_threshold,
                         prune=self.prune, prune_threshold=self.prune_threshold)
        woe.fit(x_train, y_train)
        x_woe = woe.transform(x_train)
        x_oot_woe = woe.transform(x_oot)

        # 变量稳定性
        train_bin = woe.cal_bin_ks(x_train, y_train)
        oot_bin = woe.cal_bin_ks(x_oot, y_oot, oot=True)

        # KS
        ks_ins = []
        ks_oot = []
        ks_index = []
        for i in x_woe.columns:
            try:
                ks_ins.append(max(train_bin[i][train_bin[i].ks.notnull()].ks))
                ks_oot.append(max(oot_bin[i][oot_bin[i].ks.notnull()].ks))
                ks_index.append(i)
            except Exception:
                pass

        # IV
        iv_data = []
        iv_index = []
        for i in x_woe.columns:
            try:
                tmp = sum(woe.data_matrix[i].iv)
                if tmp == 0:
                    pass
                else:
                    iv_index.append(i)
                    iv_data.append(tmp)
            except Exception:
                pass

        # PSI
        psi = pd.DataFrame(woe.cal_psi(train_bin, oot_bin), index=['psi']).T.reset_index().rename(
            columns={'index': '字段名称'})
        iv = pd.DataFrame({'字段名称': iv_index, 'Information Value': iv_data})
        ks = pd.DataFrame({'字段名称': ks_index, 'INS_KS': ks_ins, 'OOT_KS': ks_oot})
        psi_ks_iv = psi.merge(iv, how='inner', on='字段名称').merge(ks, how='inner', on='字段名称')

        if self.psi_threshold is None:
            pass
        else:
            if isinstance(self.psi_threshold, float):
                left_features = psi_ks_iv[psi_ks_iv['psi'] < self.psi_threshold]['字段名称'].tolist()
                psi_ks_iv = psi_ks_iv[psi_ks_iv['psi'] < self.psi_threshold]
                x_woe = x_woe[left_features]
                x_oot_woe = x_oot_woe[left_features]
            else:
                raise ValueError("psi_threshold must be 'all' or a float between (0, 1]")

        if self.iv_threshold is None:
            pass
        else:
            if isinstance(self.iv_threshold, float):
                left_features = psi_ks_iv[psi_ks_iv['Information Value'] >= self.iv_threshold]['字段名称'].tolist()
                psi_ks_iv = psi_ks_iv[psi_ks_iv['Information Value'] >= self.iv_threshold]
                x_woe = x_woe[left_features]
                x_oot_woe = x_oot_woe[left_features]
            else:
                raise ValueError("iv_threshold must be a float like 0.01")

        if gating:
            return x_woe, x_oot_woe, woe, psi_ks_iv
        else:
            return x_train, x_oot, woe, psi_ks_iv


class Tea:
    def __init__(self, useless_features, label='is_overdue', datetime_feature='create_time', split_method='oot'
                 , oot_threshold=None, oot_ratio=1/6, file_path='final_report.xlsx'):
        """

        :param useless_features: list
        :param label: str
        :param datetime_feature: str
        :param split_method: 'oos' or 'oot'
        :param oot_threshold: if split_method == 'oot'
        :param oot_ratio: less than 1
        :param file_path: file_path
        """
        self.oot_ratio = oot_ratio
        self.oot_threshold = oot_threshold
        self.sheets = dict()
        self.ct = None  # 分类变量bad rate替换类
        self.left_features = []  # 入模型的变量
        self.woe = None  # woe分箱类
        self.clf = None  # 分类模型
        self.X_train = None  # 原始训练数据
        self.X_oot = None  # 原始OOT数据
        self.y_train = None  # 训练的Y
        self.y_oot = None  # OOT的Y
        self.stacking_train = None  # 模型cv的训练概率分
        self.stacking_oot = None  # 模型cv的OOT概率分
        self.train_ts = None  # 训练数据的时间序列
        self.oot_ts = None  # OOT数据的时间序列
        self.datetime_feature = datetime_feature  # 时间列的列名
        self.split_method = split_method
        self.label = label  # y值列的列名
        self.useless_features = useless_features  # 不参与训练的列
        self.file_path = file_path  # 报告保存地址
        self.method = None  # 变量筛选的顺序列表
        self.encoders = None  # 编码器
        self.bins = None
        self.method_judge = []

    @staticmethod
    def get_describe(x):
        nu = []
        nu_ratio = []
        most_common = []
        most_common_ratio = []
        for i in x.columns:
            most_common.append(sum(x[i] == (x[i].value_counts().index[0])))
            most_common_ratio.append(sum(x[i] == (x[i].value_counts().index[0])) / x.shape[0])
            nu.append(sum(x[i].isnull()))
            nu_ratio.append(sum(x[i].isnull()) / x.shape[0])
        return nu, nu_ratio, most_common, most_common_ratio

    @staticmethod
    def _ks_curve(df, month=None):
        a = pd.DataFrame()
        a['decile'] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        a['bin'] = df.iloc[:, 0]
        a['size'] = df.iloc[:, 1]
        a['num_bad'] = df.iloc[:, 2]
        a['prop_bad'] = a['num_bad'] / sum(a['num_bad'])
        a['cum_prop_bad'] = a['prop_bad'][0]
        for i in range(1, a.shape[0]):
            a['cum_prop_bad'][i] = a['cum_prop_bad'][i - 1] + a['prop_bad'][i]
        a['num_good'] = a['size'] - a['num_bad']
        a['prop_good'] = a['num_good'] / sum(a['num_good'])
        a['cum_prop_good'] = a['prop_good'][0]
        for i in range(1, a.shape[0]):
            a['cum_prop_good'][i] = a['cum_prop_good'][i - 1] + a['prop_good'][i]
        a['bad_rate'] = round(a['num_bad'] / a['size'], 4)
        a['ks_curve'] = a['cum_prop_good'] - a['cum_prop_bad']
        a['ks_curve'] = round(a['ks_curve'], 4)
        if month is not None:
            a.insert(0, 'month', month)
        return a

    def wash(self, data, null_drop_rate=0.8, most_common_drop_rate=0.9):
        # ------------------------  STEP 1 训练测试集划分、Bad rate替换及变量初筛--------------------------------
        df = data.copy()
        df[self.datetime_feature] = pd.to_datetime(df[self.datetime_feature])

        if self.split_method == 'oot':
            if self.oot_threshold is None:
                train = df.sort_values(self.datetime_feature)[:df.shape[0] - int(df.shape[0] * self.oot_ratio)]\
                    .reset_index(drop=True)
                oot = df.sort_values(self.datetime_feature)[df.shape[0] - int(df.shape[0] * self.oot_ratio):].\
                    reset_index(drop=True)
            else:
                train = df[df[self.datetime_feature] < self.oot_threshold].reset_index(drop=True)
                oot = df[df[self.datetime_feature] >= self.oot_threshold].reset_index(drop=True)

        elif self.split_method == 'oos':
            from sklearn.model_selection import train_test_split
            train, oot, y_train, y_oot = train_test_split(df, df[self.label], test_size=0.2, random_state=12)
            train = train.reset_index(drop=True)
            oot = oot.reset_index(drop=True)
            y_train = y_train.reset_index(drop=True)
            y_oot = y_oot.reset_index(drop=True)
        else:
            raise KeyError("Wrong type...")

        X = df.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
        X_train = train.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
        y_train = train[self.label]
        X_oot = oot.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
        y_oot = oot[self.label]
        self.train_ts = train[self.datetime_feature]
        self.oot_ts = oot[self.datetime_feature]

        # ==== sheet 样本分析 ====
        sheet_sample = pd.DataFrame(
            {'时间跨度': [str(min(train[self.datetime_feature])) + '~' + str(max(train[self.datetime_feature])),
                      str(min(oot[self.datetime_feature])) + '~' + str(max(oot[self.datetime_feature]))],
             'Cnt': [X_train.shape[0], X_oot.shape[0]],
             'GoodCnt': [train.shape[0] - sum(train[self.label]), oot.shape[0] - sum(oot[self.label])],
             'BadCnt': [sum(train[self.label]), sum(oot[self.label])],
             'BadRate': [sum(y_train) / y_train.shape[0], sum(y_oot) / y_oot.shape[0]]})
        self.sheets['sheet_sample'] = sheet_sample
        # ==== sheet 变量缺失率 & 基本探索性分析 ====
        nu, nu_ratio, most_common, most_common_ratio = Tea.get_describe(X)

        # 变量初筛
        print("Preliminary screening...")
        sheet_2_tmp = pd.merge(pd.DataFrame(
            {'变量名称': list(X.columns), '空值个数': nu, '空值个数占比': nu_ratio, '最常值个数': most_common, '最常值个数占比': most_common_ratio}),
            pd.DataFrame(
                X.describe().T.reset_index()[['index', 'mean', 'std', 'min', '25%', '50%', '75%', 'max']]).rename(
                columns={'index': '变量名称'}),
            how='left', on='变量名称')
        left_features = list(set(sheet_2_tmp[sheet_2_tmp['空值个数占比'] < null_drop_rate]['变量名称']) & set(
            sheet_2_tmp[sheet_2_tmp['最常值个数占比'] < most_common_drop_rate]['变量名称']))

        X = X[left_features]
        X_train = X_train[left_features]
        X_oot = X_oot[left_features]

        nu, nu_ratio, most_common, most_common_ratio = Tea.get_describe(X)

        sheet_distribution = pd.concat([pd.DataFrame(
            {'变量名称': list(X.columns), '空值个数': nu, '空值个数占比': nu_ratio, '最常值个数': most_common, '最常值个数占比': most_common_ratio}),
            pd.DataFrame(X.describe().T.reset_index()[
                             ['mean', 'std', 'min', '25%', '50%', '75%', 'max']])], axis=1)

        self.sheets['sheet_distribution'] = sheet_distribution
        self.X_train = X_train
        self.X_oot = X_oot
        self.y_train = y_train
        self.y_oot = y_oot

    def cook(self, encoders):
        self.encoders = encoders
        for estimator in self.encoders:
            if isinstance(estimator, TeaOneHotEncoder):
                self.X_train = estimator.fit_transform(self.X_train)
                self.X_oot = estimator.transform(self.X_oot)
            else:
                self.X_train = estimator.fit_transform(self.X_train, self.y_train)
                self.X_oot = estimator.transform(self.X_oot)

    def select(self, method):
        self.method = method
        _X = self.X_train.copy()
        _X_oot = self.X_oot.copy()

        try:
            self.method_judge = [1 if isinstance(estimator, WOE) else 0 for estimator in method]
        except:
            pass

        if sum(self.method_judge) != 0:
            for estimator in self.method:
                if isinstance(estimator, WOE):
                    _X, _X_oot, woe, psi_ks_iv = estimator.woe_processing(_X, self.y_train, _X_oot, self.y_oot, gating=True)
                elif isinstance(estimator, OutlierTransform) \
                        or isinstance(estimator, FilterCorr) \
                        or isinstance(estimator, FilterVif):
                    _X = estimator.fit_transform(_X)
                    _X_oot = estimator.transform(_X_oot)
                else:
                    _X = estimator.fit_transform(_X, self.y_train)
                    _X_oot = estimator.transform(_X_oot)
        elif sum(self.method_judge) == 0:
            for estimator in self.method:
                if isinstance(estimator, OutlierTransform) \
                        or isinstance(estimator, FilterCorr) \
                        or isinstance(estimator, FilterVif):
                    _X = estimator.fit_transform(_X)
                    _X_oot = estimator.transform(_X_oot)
                else:
                    _X = estimator.fit_transform(_X, self.y_train)
                    _X_oot = estimator.transform(_X_oot)
            wow = WOE(bins=10, bad_rate_merge=True, bad_rate_sim_threshold=0.05)
            _X, _X_oot, woe, psi_ks_iv = wow.woe_processing(_X, self.y_train, _X_oot, self.y_oot, gating=False)

        X_woe, X_oot_woe = _X, _X_oot
        self.left_features = X_woe.columns.tolist()
        sheet_psi_ks_iv = psi_ks_iv[psi_ks_iv['字段名称'].isin(self.left_features)].reset_index(drop=True)
        self.sheets['sheet_psi_ks_iv'] = sheet_psi_ks_iv
        self.woe = woe
        self.X_woe = X_woe
        self.X_oot_woe = X_oot_woe

    def drink(self, clf):
        if self.method_judge != 0:
            self.X_oot_woe = self.X_oot_woe.fillna(0.0)
        else:
            pass
        ssss = StratifiedKFold(n_splits=5, random_state=11, shuffle=True)
        # clf = LogisticRegression(penalty='l2', C=0.5)
        clf, stacking_train, stacking_oot = train_by_cv(self.X_woe, self.y_train, self.X_oot_woe, self.y_oot, ssss, clf)
        clf.fit(self.X_woe, self.y_train)
        self.clf = clf
        self.stacking_train = stacking_train
        self.stacking_oot = stacking_oot

    def sleep(self, bins):
        # ------------------------------  STEP 4 整理表  ---------------------------------------------------------------
        # ==== sheet 各个bin的变量逾期分布和KS值 ====
        self.bins = bins
        train_bin = self.woe.cal_bin_ks(self.X_train[self.left_features], self.y_train)
        oot_bin = self.woe.cal_bin_ks(self.X_oot[self.left_features], self.y_oot, oot=True)

        # 评分相关性
        sheet_correlations = self.X_train[self.left_features].corr()

        # ==== sheet 特征权重和预测概率分析 ====
        # result_train = pd.concat([X_woe, pd.DataFrame(stacking_train)], axis=1).rename(columns={0: 'model_score'})
        result_test = pd.concat([self.X_oot_woe, pd.DataFrame(self.stacking_oot)], axis=1).rename(
            columns={0: 'model_score'})
        im_final = get_importance(self.clf, self.X_woe).reset_index(drop=True)

        model_col = list(set(result_test.columns) - {'model_score'})
        score_corr = []
        for i in model_col:
            score_corr.append(result_test[i].corr(result_test['model_score']))

        label_corr = []
        for i in model_col:
            label_corr.append(result_test[i].corr(pd.Series(self.y_oot)))

        tmp = pd.merge(pd.DataFrame({'feature_name': model_col, '与预测概率相关性': score_corr}),
                       pd.DataFrame({'feature_name': model_col, '与风险相关方向': label_corr}),
                       how='inner',
                       on='feature_name')

        # ==== sheet 模型分析，ks曲线，累计bad rate... ====
        sheet_weights = pd.merge(tmp, im_final, how='inner', on='feature_name')
        model_info_tmp_train = pd.concat(
            [self.train_ts, self.y_train, pd.DataFrame(self.stacking_train)],
            axis=1).rename(columns={0: 'model_score'})
        model_info_tmp_oot = pd.concat(
            [self.oot_ts, self.y_oot, pd.DataFrame(self.stacking_oot)],
            axis=1).rename(columns={0: 'model_score'})
        model_info_tmp_train['tag'] = 'INS'
        model_info_tmp_oot['tag'] = 'OOT'
        model_info_tmp = pd.concat([model_info_tmp_train, model_info_tmp_oot])
        model_info_tmp['bin'] = pd.qcut(model_info_tmp['model_score'], 10)
        model_info_tmp['month'] = model_info_tmp[self.datetime_feature].dt.strftime('%Y%m')

        model_info_tmp_ins = pd.DataFrame(model_info_tmp[model_info_tmp['tag'] == 'INS'].groupby('bin').agg(
            {self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
        model_info_tmp_oot = pd.DataFrame(model_info_tmp[model_info_tmp['tag'] == 'OOT'].groupby('bin').agg(
            {self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
        sheet_model_info_ins = Tea._ks_curve(model_info_tmp_ins)
        sheet_model_info_ins['tag'] = 'INS'
        sheet_model_info_oot = Tea._ks_curve(model_info_tmp_oot)
        sheet_model_info_oot['tag'] = 'OOT'

        # -------------------------------  STEP 5 写入表  ---------------------------------------------------------------
        bin_ks_ins = pd.DataFrame()
        bin_ks_oot = pd.DataFrame()
        for i in train_bin.keys():
            train_tmp = train_bin[i].reset_index().rename(columns={'index': 'bins'})
            train_tmp.insert(0, 'feature', i)
            bin_ks_ins = pd.concat([bin_ks_ins, train_tmp])
            oot_tmp = oot_bin[i].reset_index().rename(columns={'index': 'bins'})
            oot_tmp.insert(0, 'feature', i)
            bin_ks_oot = pd.concat([bin_ks_oot, oot_tmp])

        self.sheets['sheet_feature_bin_ins'] = bin_ks_ins
        self.sheets['sheet_feature_bin_oot'] = bin_ks_oot
        self.sheets['sheet_correlations'] = sheet_correlations
        self.sheets['sheet_weights'] = sheet_weights
        self.sheets['sheet_model_info_ins'] = sheet_model_info_ins
        self.sheets['sheet_model_info_oot'] = sheet_model_info_oot

        writer = pd.ExcelWriter(self.file_path)
        self.sheets['sheet_sample'].to_excel(writer, sheet_name='样本分析', index=False)
        self.sheets['sheet_distribution'].to_excel(writer, sheet_name='变量缺失率 & 基本探索性分析', index=False)
        self.sheets['sheet_psi_ks_iv'].to_excel(writer, sheet_name='INS变量IV值 & 时间稳定性', index=False)

        row_index = 0
        for i in train_bin.keys():
            train_tmp = train_bin[i].reset_index().rename(columns={'index': i})
            train_tmp['bad'] = '[' + train_tmp['left'].astype(str) + ',' + train_tmp['right'].astype(str) + ')'
            oot_tmp = oot_bin[i].reset_index().rename(columns={'index': i})
            oot_tmp['bad'] = '[' + oot_tmp['left'].astype(str) + ',' + oot_tmp['right'].astype(str) + ')'
            train_tmp.to_excel(writer, startrow=row_index, startcol=1, sheet_name='变量逾期分布和KS值', index=False)
            oot_tmp.to_excel(writer, startrow=row_index, startcol=13, sheet_name='变量逾期分布和KS值', index=False)
            row_index += self.bins+2

        sheet_correlations.to_excel(writer, sheet_name='评分相关性', index=True)
        sheet_weights.to_excel(writer, sheet_name='模型', index=False)
        sheet_model_info_ins.to_excel(writer, sheet_name='模型', startcol=6, index=False)
        sheet_model_info_oot.to_excel(writer, sheet_name='模型', startrow=11, startcol=6, index=False, header=False)

        row_index_8 = 0
        for month in model_info_tmp['month'].unique():
            trace_back = model_info_tmp[model_info_tmp['month'] == month].drop(['month'], axis=1).reset_index(drop=True)
            trace_back = pd.DataFrame(
                trace_back.groupby('bin').agg({self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
            sheet_trace_back = Tea._ks_curve(trace_back, month)
            if row_index_8 == 0:
                sheet_trace_back.to_excel(writer, sheet_name='模型回测', startrow=row_index_8, startcol=0, index=False)
                row_index_8 += 11
            else:
                sheet_trace_back.to_excel(writer, sheet_name='模型回测', startrow=row_index_8, startcol=0, index=False,
                                          header=False)
                row_index_8 += 10
        writer.save()

        # -------------------------------  STEP 6 美化（字体/字号/边框/颜色/粗细）  --------------------------------------------
        wb = openpyxl.load_workbook(self.file_path)
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4

        sheet = wb['样本分析']
        for i in sheet['A1':'E3']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['A1':'E1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        sheet = wb['变量缺失率 & 基本探索性分析']
        for i in sheet['A1':'L%s' % (self.sheets['sheet_distribution'].shape[0] + 1)]:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['A1':'L1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        sheet = wb['INS变量IV值 & 时间稳定性']
        for i in sheet['A1':'E%s' % (self.sheets['sheet_psi_ks_iv'].shape[0] + 1)]:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['A1':'E1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        sheet = wb['变量逾期分布和KS值']
        for i in sheet['A1':'X%s' % row_index]:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].fill = eval(conf.get('config', 'fill2'))
        for ind in range(1, row_index, self.bins+2):
            for i in sheet['B%s' % ind:'L%s' % ind]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            for i in sheet['N%s' % ind:'X%s' % ind]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            sheet['A%s' % ind] = 'INS'
            sheet['M%s' % ind] = 'OOT'
            sheet['A%s' % ind].font = eval(conf.get('config', 'font3'))
            sheet['M%s' % ind].font = eval(conf.get('config', 'font3'))

        for i in sheet['A1':'A%s' % row_index]:
            for j in range(len(i)):
                i[j].fill = eval(conf.get('config', 'fill3'))

        for i in sheet['M1':'M%s' % row_index]:
            for j in range(len(i)):
                i[j].fill = eval(conf.get('config', 'fill3'))

        sheet = wb['模型']
        for i in sheet['A1':'D%s' % (sheet_weights.shape[0] + 1)]:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['A1':'D1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        for i in sheet['G1':'R21']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['G1':'R1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        cnt = 2
        while sheet['R'+str(cnt)].value is not None:
            c1 = BarChart()
            v1 = Reference(sheet, min_col=16, min_row=cnt-1, max_col=16, max_row=cnt+9)
            c1.add_data(v1, titles_from_data=True)

            c1.x_axis.title = 'Bin Decile'
            c1.y_axis.title = 'Bad Rate'
            c1.y_axis.majorGridlines = None
            c1.title = sheet['R'+str(cnt)].value

            # Create a second chart
            c2 = LineChart()
            v2 = Reference(sheet, min_col=17, min_row=cnt-1, max_col=17, max_row=cnt+9)
            c2.add_data(v2, titles_from_data=True)
            c2.y_axis.axId = 200
            c2.y_axis.title = "Ks"

            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            c1.y_axis.crosses = "max"
            c1 += c2

            sheet.add_chart(c1, 'T'+str(cnt))
            cnt += 10

        sheet = wb['模型回测']
        for i in sheet['A1':'L%s' % row_index_8]:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font1'))
                i[j].alignment = eval(conf.get('config', 'alignment1'))
                i[j].border = eval(conf.get('config', 'border1'))
        for i in sheet['A1':'L1']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))

        cnt = 2
        while sheet['A'+str(cnt)].value is not None:
            c1 = BarChart()
            v1 = Reference(sheet, min_col=11, min_row=cnt-1, max_col=11, max_row=cnt+9)
            c1.add_data(v1, titles_from_data=True)

            c1.x_axis.title = 'Bin Decile'
            c1.y_axis.title = 'Bad Rate'
            c1.y_axis.majorGridlines = None
            c1.title = sheet['A'+str(cnt)].value

            # Create a second chart
            c2 = LineChart()
            v2 = Reference(sheet, min_col=12, min_row=cnt-1, max_col=12, max_row=cnt+9)
            c2.add_data(v2, titles_from_data=True)
            c2.y_axis.axId = 200
            c2.y_axis.title = "Ks"

            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            c1.y_axis.crosses = "max"
            c1 += c2

            sheet.add_chart(c1, 'N'+str(cnt))
            cnt += 10

        wb.create_sheet(title='数据测试说明', index=0)
        head = wb['数据测试说明']
        head['A1'] = '数据测试说明'
        head.merge_cells('A1:B1')
        head['A2'] = '提供的样本'
        head['A3'] = '测试数据变量类型'
        head['A4'] = '测试内容'
        head.merge_cells('A4:A9')
        head['A10'] = '测试结论'
        head['B4'] = '样本分析'
        head['B5'] = '变量缺失率 & 基本探索性分析'
        head['B6'] = 'INS变量IV值 & 时间稳定性'
        head['B7'] = '变量逾期分布和KS值'
        head['B8'] = '评分相关性'
        head['B9'] = '模型'
        for i in head['A1':'B10']:
            for j in range(len(i)):
                if str(i[j])[15:18] in ('B4>', 'B5>', 'B6>', 'B7>', 'B8>', 'B9'):
                    i[j].font = eval(conf.get('config', 'font4'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
                else:
                    i[j].font = eval(conf.get('config', 'font4'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
        for i in head['A2':'A10']:
            for j in range(len(i)):
                i[j].font = eval(conf.get('config', 'font2'))
                i[j].fill = eval(conf.get('config', 'fill1'))
        for i in ['A1', 'B1']:
            head[i].font = eval(conf.get('config', 'font2'))
            head[i].fill = eval(conf.get('config', 'fill1'))

        wb.save(self.file_path)
        print('Finish 🍵 ')
